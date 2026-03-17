from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from setup import XLSX_PATH

class excel_creater:
    def __init__(self):
        pass

    def append_row(
        self,
        session: str,
        theory: str,
        attempt: int,
        error_message: str,
        fixes_text: str,
        status: str,  # "success" or "fail"
        elapsed_seconds: float | None = None,
        sheet_name: str = "llm-repair",
    ) -> None:
        """
        Append one attempt row to an Excel file.

        Columns (exactly):
          1) session+theory
          2) attempt
          3) error message
          4) fixes
          5) fail/success
          6) elapsed_seconds (total time for the theory on the final attempt; blank otherwise)
        """
        xlsx_path = Path(XLSX_PATH)
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if xlsx_path.exists():
            wb = openpyxl.load_workbook(xlsx_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

        required_headers = [
            "session+theory",
            "attempt",
            "error message",
            "fixes",
            "fail/success",
            "elapsed_seconds",
        ]
        column_widths = {
            "session+theory": 40,
            "attempt": 10,
            "error message": 80,
            "fixes": 80,
            "fail/success": 12,
            "elapsed_seconds": 16,
        }

        headers = [cell.value for cell in ws[1]]
        if not headers or all(h is None for h in headers):
            headers = []

        def normalize_header(value: object) -> str:
            return str(value).strip().lower()

        def ensure_header(name: str) -> int:
            for idx, value in enumerate(headers, start=1):
                if value is None:
                    continue
                if normalize_header(value) == normalize_header(name):
                    return idx
            new_idx = len(headers) + 1
            headers.append(name)
            cell = ws.cell(row=1, column=new_idx, value=name)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border
            width = column_widths.get(name)
            if width:
                ws.column_dimensions[get_column_letter(new_idx)].width = width
            return new_idx

        col_idx = {name: ensure_header(name) for name in required_headers}

        key = f"{session}/{theory}"
        row = ws.max_row + 1
        values = {
            "session+theory": key,
            "attempt": attempt,
            "error message": error_message,
            "fixes": fixes_text,
            "fail/success": status,
            "elapsed_seconds": (round(elapsed_seconds, 3) if elapsed_seconds is not None else None),
        }

        # wrap text for the newly appended row
        for header, value in values.items():
            c = col_idx[header]
            cell = ws.cell(row=row, column=c, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border

        wb.save(xlsx_path)
